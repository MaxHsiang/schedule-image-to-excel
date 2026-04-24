# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List, Sequence, Tuple

from openpyxl import load_workbook

from schedule_core import ShiftRecord, export_to_excel, timestamped_output_path


SHIFT_ROWS: Sequence[Tuple[str, int]] = (("早", 1), ("午", 2), ("晚", 3))
SHIFT_HOURS = {"早": 4, "午": 5, "晚": 4}
NAME_VARIANTS = {
    "张盈慧": "張盈慧",
    "張盈慧": "張盈慧",
}
WEEKDAY_MAP = "一二三四五六日"


def normalize_name(text: str) -> str:
    normalized = str(text).strip().replace(" ", "")
    return NAME_VARIANTS.get(normalized, normalized)


def parse_schedule_excel(excel_path: Path, employee_name: str) -> tuple[list[ShiftRecord], dict]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    target_name = normalize_name(employee_name)
    records: List[ShiftRecord] = []

    for date_row in range(3, sheet.max_row + 1, 4):
        if date_row + 3 > sheet.max_row:
            continue

        for col in range(2, 9):
            work_date = sheet.cell(date_row, col).value
            if not isinstance(work_date, datetime):
                continue

            for shift, row_offset in SHIFT_ROWS:
                value = sheet.cell(date_row + row_offset, col).value
                if value is None:
                    continue
                if normalize_name(value) != target_name:
                    continue

                records.append(
                    ShiftRecord(
                        year=work_date.year,
                        month=work_date.month,
                        day=work_date.day,
                        weekday=WEEKDAY_MAP[work_date.weekday()],
                        shift=shift,
                        hours=SHIFT_HOURS[shift],
                    )
                )

    records.sort(key=lambda item: (item.year, item.month, item.day, ("早", "午", "晚").index(item.shift)))
    debug = {
        "source": "excel",
        "sheet": sheet.title,
        "total_rows": sheet.max_row,
        "matched_records": len(records),
    }
    return records, debug


def run_excel_conversion_debug(
    excel_path: Path,
    employee_name: str,
    output_path: Path | None = None,
) -> tuple[Path, list[ShiftRecord], dict]:
    records, debug = parse_schedule_excel(excel_path, employee_name)
    final_output = output_path or timestamped_output_path(Path(excel_path), employee_name, records)
    saved_path = export_to_excel(records=records, employee_name=employee_name, output_path=final_output)
    return saved_path, records, debug
