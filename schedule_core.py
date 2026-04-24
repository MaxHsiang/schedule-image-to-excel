# -*- coding: utf-8 -*-
from __future__ import annotations

import re
from bisect import bisect_right
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import cv2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rapidocr_onnxruntime import RapidOCR


FONT_NAME = "Microsoft JhengHei"
SHIFT_HOURS = {"早": 4, "午": 5, "晚": 4}
SHIFT_ORDER = {"早": 0, "午": 1, "晚": 2}
SHIFT_ROW_FILLS = {
    "早": PatternFill("solid", fgColor="EAF8FF"),
    "午": PatternFill("solid", fgColor="FFF2D9"),
    "晚": PatternFill("solid", fgColor="EEF8E3"),
}
SHIFT_CELL_FILLS = {
    "早": PatternFill("solid", fgColor="BFEFFF"),
    "午": PatternFill("solid", fgColor="FFD27F"),
    "晚": PatternFill("solid", fgColor="9AD97A"),
}
WEEKDAY_MAP = "一二三四五六日"
NAME_VARIANTS = {
    "張盈慧": "張盈慧",
    "张盈慧": "張盈慧",
}


@dataclass
class ShiftRecord:
    year: int
    month: int
    day: int
    weekday: str
    shift: str
    hours: int


class ScheduleParser:
    def __init__(self) -> None:
        self.ocr = RapidOCR()

    def parse(self, image_path: Path, employee_name: str) -> List[ShiftRecord]:
        image_path = Path(image_path)
        image = cv2.imread(str(image_path))
        if image is None:
            raise ValueError("圖片讀取失敗。")

        ocr_results = self._run_ocr(image_path)
        year, month = self._extract_year_month(ocr_results)
        verticals, horizontals = self._detect_grid_lines_from_image(image)
        date_cells = self._collect_dates(ocr_results, year, month, verticals, horizontals)
        return self._collect_records(
            image=image,
            ocr_results=ocr_results,
            employee_name=employee_name,
            date_cells=date_cells,
            verticals=verticals,
            horizontals=horizontals,
        )

    def _run_ocr(self, image_path: Path):
        results, _ = self.ocr(str(image_path))
        if not results:
            raise ValueError("OCR 沒有讀到任何內容，請確認圖片清晰且完整。")
        return results

    def _parse_ocr_item(self, item) -> Tuple[Sequence[Sequence[float]], str, float]:
        if not isinstance(item, (list, tuple)) or len(item) < 3:
            raise ValueError("OCR 回傳格式不正確。")

        first, second, third = item[0], item[1], item[2]
        if isinstance(second, str):
            box = first
            text = second
            score_raw = third
        else:
            box = first
            text = third
            score_raw = second

        try:
            score = float(score_raw)
        except Exception:
            score = 0.0

        return box, str(text), score

    def _extract_year_month(self, ocr_results) -> Tuple[int, int]:
        for item in ocr_results:
            _, text, _ = self._parse_ocr_item(item)
            match = re.search(r"(\d{4})\s*/\s*(\d{1,2})", self._normalize_text(text))
            if match:
                return int(match.group(1)), int(match.group(2))
        raise ValueError("無法從圖片標題辨識出年份與月份。")

    def _detect_grid_lines_from_image(self, image) -> Tuple[List[int], List[int]]:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        dark_mask = gray < 180
        col_counts = dark_mask.sum(axis=0)
        row_counts = dark_mask.sum(axis=1)

        verticals = self._group_lines(
            [idx for idx, count in enumerate(col_counts) if count > gray.shape[0] * 0.7]
        )
        horizontals = self._group_lines(
            [idx for idx, count in enumerate(row_counts) if count > gray.shape[1] * 0.7]
        )

        verticals = self._with_edges(verticals, gray.shape[1] - 1)
        horizontals = self._with_edges(horizontals, gray.shape[0] - 1)

        if len(verticals) < 9 or len(horizontals) < 10:
            raise ValueError("無法穩定抓到班表格線，請使用完整正拍的班表圖片。")
        return verticals, horizontals

    def _group_lines(self, values: Sequence[int], gap: int = 2) -> List[int]:
        groups: List[List[int]] = []
        for value in values:
            if not groups or value - groups[-1][-1] > gap:
                groups.append([value])
            else:
                groups[-1].append(value)
        return [round(sum(group) / len(group)) for group in groups]

    def _with_edges(self, lines: List[int], max_index: int) -> List[int]:
        boundaries = list(lines)
        if not boundaries or boundaries[0] > 8:
            boundaries = [0] + boundaries
        else:
            boundaries[0] = 0

        if boundaries[-1] < max_index - 8:
            boundaries.append(max_index)
        else:
            boundaries[-1] = max_index
        return boundaries

    def _collect_dates(
        self,
        ocr_results,
        year: int,
        month: int,
        verticals: Sequence[int],
        horizontals: Sequence[int],
    ) -> Dict[Tuple[int, int], date]:
        date_cells: Dict[Tuple[int, int], date] = {}

        for item in ocr_results:
            box, text, score = self._parse_ocr_item(item)
            if score < 0.7:
                continue

            normalized = self._normalize_text(text)
            match = re.fullmatch(r"(\d{2})/(\d{2})", normalized)
            if not match:
                continue

            cell_col, cell_row = self._locate_cell(box, verticals, horizontals)
            if cell_col < 1 or cell_col > 7:
                continue

            mm = int(match.group(1))
            dd = int(match.group(2))
            if mm != month:
                continue

            date_cells[(cell_row, cell_col)] = date(year, month, dd)

        if not date_cells:
            raise ValueError("沒有辨識到任何日期欄位。")
        return date_cells

    def _collect_records(
        self,
        image,
        ocr_results,
        employee_name: str,
        date_cells: Dict[Tuple[int, int], date],
        verticals: Sequence[int],
        horizontals: Sequence[int],
    ) -> List[ShiftRecord]:
        target_name = self._normalize_name(employee_name)
        records: List[ShiftRecord] = []
        matched_cells = set()
        occupied_cells: Dict[Tuple[int, int], str] = {}
        target_samples: List[Tuple[float, float, float]] = []

        for item in ocr_results:
            box, text, score = self._parse_ocr_item(item)
            if score < 0.72:
                continue

            cell_col, cell_row = self._locate_cell(box, verticals, horizontals)
            if cell_col < 1 or cell_col > 7 or cell_row < 2:
                continue

            shift_index = (cell_row - 2) % 4
            if shift_index not in (0, 1, 2):
                continue

            normalized_name = self._normalize_name(text)
            occupied_cells[(cell_row, cell_col)] = normalized_name
            if normalized_name != target_name:
                continue

            date_row = cell_row - shift_index - 1
            work_date = date_cells.get((date_row, cell_col))
            if work_date is None:
                continue

            shift = ("早", "午", "晚")[shift_index]
            records.append(
                ShiftRecord(
                    year=work_date.year,
                    month=work_date.month,
                    day=work_date.day,
                    weekday=WEEKDAY_MAP[work_date.weekday()],
                    shift=shift,
                    hours=SHIFT_HOURS[shift],
                )
            )
            matched_cells.add((cell_row, cell_col))

            sample = self._cell_mean_bgr(image, cell_row, cell_col, verticals, horizontals)
            if sample is not None:
                target_samples.append(sample)

        if target_samples:
            prototype = tuple(sum(v) / len(v) for v in zip(*target_samples))
            for (date_row, cell_col), work_date in date_cells.items():
                for offset, shift in enumerate(("早", "午", "晚"), start=1):
                    cell_row = date_row + offset
                    key = (cell_row, cell_col)

                    if key in matched_cells:
                        continue
                    if key in occupied_cells and occupied_cells[key] != target_name:
                        continue

                    sample = self._cell_mean_bgr(image, cell_row, cell_col, verticals, horizontals)
                    if sample is None:
                        continue
                    if self._is_target_color(sample, prototype):
                        records.append(
                            ShiftRecord(
                                year=work_date.year,
                                month=work_date.month,
                                day=work_date.day,
                                weekday=WEEKDAY_MAP[work_date.weekday()],
                                shift=shift,
                                hours=SHIFT_HOURS[shift],
                            )
                        )
                        matched_cells.add(key)

        if not records:
            raise ValueError(f"找不到 {employee_name} 的班次。")

        dedup = {}
        for record in records:
            dedup[(record.year, record.month, record.day, record.shift)] = record

        final_records = list(dedup.values())
        final_records.sort(key=lambda item: (item.year, item.month, item.day, SHIFT_ORDER[item.shift]))
        return final_records

    def _cell_mean_bgr(
        self,
        image,
        cell_row: int,
        cell_col: int,
        verticals: Sequence[int],
        horizontals: Sequence[int],
    ) -> Tuple[float, float, float] | None:
        if cell_col + 1 >= len(verticals) or cell_row + 1 >= len(horizontals):
            return None

        x1, x2 = int(verticals[cell_col]), int(verticals[cell_col + 1])
        y1, y2 = int(horizontals[cell_row]), int(horizontals[cell_row + 1])
        pad_x = max(4, (x2 - x1) // 6)
        pad_y = max(4, (y2 - y1) // 6)
        roi = image[y1 + pad_y : y2 - pad_y, x1 + pad_x : x2 - pad_x]
        if roi.size == 0:
            return None
        mean = roi.mean(axis=(0, 1))
        return float(mean[0]), float(mean[1]), float(mean[2])

    def _is_target_color(
        self,
        sample_bgr: Tuple[float, float, float],
        target_bgr: Tuple[float, float, float],
    ) -> bool:
        sb, sg, sr = sample_bgr
        tb, tg, tr = target_bgr
        distance = ((sb - tb) ** 2 + (sg - tg) ** 2 + (sr - tr) ** 2) ** 0.5
        brightness = (sb + sg + sr) / 3
        green_bias = sg - max(sb, sr)
        return distance < 65 and brightness < 235 and green_bias > 8

    def _locate_cell(
        self,
        box: Sequence[Sequence[float]],
        verticals: Sequence[int],
        horizontals: Sequence[int],
    ) -> Tuple[int, int]:
        xs = [point[0] for point in box]
        ys = [point[1] for point in box]
        center_x = sum(xs) / len(xs)
        center_y = sum(ys) / len(ys)
        cell_col = bisect_right(verticals, center_x) - 1
        cell_row = bisect_right(horizontals, center_y) - 1
        return cell_col, cell_row

    def _normalize_text(self, text: str) -> str:
        return (
            text.strip()
            .replace("（", "(")
            .replace("）", ")")
            .replace(" ", "")
            .replace("月份", "/")
            .replace("排班表", "")
        )

    def _normalize_name(self, text: str) -> str:
        normalized = self._normalize_text(text)
        return NAME_VARIANTS.get(normalized, normalized)


def export_to_excel(records: Sequence[ShiftRecord], employee_name: str, output_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "個人班表"

    headers = ["年", "月", "日", "周幾", "班別", "時數", "時薪", "單量", "總薪水"]
    sheet.append(headers)

    for record in records:
        sheet.append(
            [
                record.year,
                record.month,
                record.day,
                record.weekday,
                record.shift,
                record.hours,
                "",
                "",
                "",
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = None

    title_fill = PatternFill("solid", fgColor="2F6B53")
    header_fill = PatternFill("solid", fgColor="2F6B53")
    border = Border(
        left=Side(style="thin", color="D9E2DC"),
        right=Side(style="thin", color="D9E2DC"),
        top=Side(style="thin", color="D9E2DC"),
        bottom=Side(style="thin", color="D9E2DC"),
    )

    base_font = Font(name=FONT_NAME, size=11)
    title_font = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
    header_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    shift_font = Font(name=FONT_NAME, size=11, bold=True, color="2A3B33")

    title = f"{employee_name} 個人班表"
    sheet.insert_rows(1)
    sheet.merge_cells("A1:I1")
    sheet["A1"] = title
    sheet["A1"].font = title_font
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24

    for cell in sheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_index in range(3, sheet.max_row + 1):
        shift_value = sheet[f"E{row_index}"].value
        row_fill = SHIFT_ROW_FILLS.get(shift_value)
        for cell in sheet[row_index]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = base_font
            if row_fill is not None:
                cell.fill = row_fill

        shift_cell = sheet[f"E{row_index}"]
        if shift_value in SHIFT_CELL_FILLS:
            shift_cell.fill = SHIFT_CELL_FILLS[shift_value]
            shift_cell.font = shift_font

    widths = {"A": 10, "B": 8, "C": 8, "D": 8, "E": 10, "F": 8, "G": 12, "H": 12, "I": 14}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def default_output_path(image_path: Path, employee_name: str) -> Path:
    safe_name = employee_name.replace("/", "_").replace("\\", "_")
    return image_path.with_name(f"{image_path.stem}_{safe_name}_個人班表.xlsx")


def run_conversion(
    image_path: Path,
    employee_name: str,
    output_path: Path | None = None,
) -> Tuple[Path, List[ShiftRecord]]:
    parser = ScheduleParser()
    records = parser.parse(image_path=image_path, employee_name=employee_name)
    final_output = output_path or default_output_path(Path(image_path), employee_name)
    saved_path = export_to_excel(records=records, employee_name=employee_name, output_path=final_output)
    return saved_path, records
