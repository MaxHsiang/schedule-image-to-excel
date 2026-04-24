# -*- coding: utf-8 -*-
import argparse
import re
import traceback
from bisect import bisect_right
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import cv2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from rapidocr_onnxruntime import RapidOCR

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception:
    tk = None
    filedialog = None
    messagebox = None
    ttk = None


FONT_NAME = "Microsoft JhengHei"
SHIFT_HOURS = {"早": 4, "午": 5, "晚": 4}
SHIFT_ORDER = {"早": 0, "午": 1, "晚": 2}
SHIFT_FILLS = {
    "早": PatternFill("solid", fgColor="BFEFFF"),
    "午": PatternFill("solid", fgColor="FFD27F"),
    "晚": PatternFill("solid", fgColor="9AD97A"),
}
WEEKDAY_MAP = "一二三四五六日"
NAME_VARIANTS = {
    "张盈慧": "張盈慧",
    "張盈慧": "張盈慧",
}


@dataclass
class ShiftRecord:
    year: int
    month: int
    day: int
    weekday: str
    shift: str
    hours: int


class ScheduleParser:
    def __init__(self) -> None:
        self.ocr = RapidOCR()

    def parse(self, image_path: Path, employee_name: str) -> List[ShiftRecord]:
        image_path = Path(image_path)
        ocr_results = self._run_ocr(image_path)
        year, month = self._extract_year_month(ocr_results)
        verticals, horizontals = self._detect_grid_lines(image_path)
        date_cells = self._collect_dates(ocr_results, year, month, verticals, horizontals)
        return self._collect_records(
            ocr_results=ocr_results,
            employee_name=employee_name,
            date_cells=date_cells,
            verticals=verticals,
            horizontals=horizontals,
        )

    def _run_ocr(self, image_path: Path):
        results, _ = self.ocr(str(image_path))
        if not results:
            raise ValueError("OCR 沒有讀到任何內容，請確認圖片清晰且完整。")
        return results

    def _extract_year_month(self, ocr_results) -> Tuple[int, int]:
        for _, text, _ in ocr_results:
            match = re.search(r"(\d{4})\s*/\s*(\d{1,2})", self._normalize_text(text))
            if match:
                return int(match.group(1)), int(match.group(2))
        raise ValueError("無法從圖片標題辨識出年份與月份。")

    def _detect_grid_lines(self, image_path: Path) -> Tuple[List[int], List[int]]:
        image = cv2.imread(str(image_path))
        if image is None:
            raise ValueError("圖片讀取失敗。")

        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        dark_mask = gray < 180
        col_counts = dark_mask.sum(axis=0)
        row_counts = dark_mask.sum(axis=1)

        verticals = self._group_lines(
            [idx for idx, count in enumerate(col_counts) if count > gray.shape[0] * 0.7]
        )
        horizontals = self._group_lines(
            [idx for idx, count in enumerate(row_counts) if count > gray.shape[1] * 0.7]
        )

        verticals = self._with_edges(verticals, gray.shape[1] - 1)
        horizontals = self._with_edges(horizontals, gray.shape[0] - 1)

        if len(verticals) < 9 or len(horizontals) < 10:
            raise ValueError("無法穩定抓到班表格線，請使用完整正拍的班表圖片。")

        return verticals, horizontals

    def _group_lines(self, values: Sequence[int], gap: int = 2) -> List[int]:
        groups: List[List[int]] = []
        for value in values:
            if not groups or value - groups[-1][-1] > gap:
                groups.append([value])
            else:
                groups[-1].append(value)
        return [round(sum(group) / len(group)) for group in groups]

    def _with_edges(self, lines: List[int], max_index: int) -> List[int]:
        boundaries = list(lines)
        if not boundaries or boundaries[0] > 8:
            boundaries = [0] + boundaries
        else:
            boundaries[0] = 0

        if boundaries[-1] < max_index - 8:
            boundaries.append(max_index)
        else:
            boundaries[-1] = max_index
        return boundaries

    def _collect_dates(
        self,
        ocr_results,
        year: int,
        month: int,
        verticals: Sequence[int],
        horizontals: Sequence[int],
    ) -> Dict[Tuple[int, int], date]:
        date_cells: Dict[Tuple[int, int], date] = {}

        for box, text, score in ocr_results:
            if score < 0.7:
                continue

            normalized = self._normalize_text(text)
            match = re.fullmatch(r"(\d{2})/(\d{2})", normalized)
            if not match:
                continue

            cell_col, cell_row = self._locate_cell(box, verticals, horizontals)
            if cell_col < 1 or cell_col > 7:
                continue

            mm = int(match.group(1))
            dd = int(match.group(2))
            if mm != month:
                continue

            date_cells[(cell_row, cell_col)] = date(year, month, dd)

        if not date_cells:
            raise ValueError("沒有辨識到任何日期欄位。")
        return date_cells

    def _collect_records(
        self,
        ocr_results,
        employee_name: str,
        date_cells: Dict[Tuple[int, int], date],
        verticals: Sequence[int],
        horizontals: Sequence[int],
    ) -> List[ShiftRecord]:
        target_name = self._normalize_name(employee_name)
        records: List[ShiftRecord] = []

        for box, text, score in ocr_results:
            if score < 0.72:
                continue

            normalized_name = self._normalize_name(text)
            if normalized_name != target_name:
                continue

            cell_col, cell_row = self._locate_cell(box, verticals, horizontals)
            if cell_col < 1 or cell_col > 7 or cell_row < 2:
                continue

            shift_index = (cell_row - 2) % 4
            if shift_index not in (0, 1, 2):
                continue

            date_row = cell_row - shift_index - 1
            work_date = date_cells.get((date_row, cell_col))
            if work_date is None:
                continue

            shift = ("早", "午", "晚")[shift_index]
            records.append(
                ShiftRecord(
                    year=work_date.year,
                    month=work_date.month,
                    day=work_date.day,
                    weekday=WEEKDAY_MAP[work_date.weekday()],
                    shift=shift,
                    hours=SHIFT_HOURS[shift],
                )
            )

        if not records:
            raise ValueError(f"找不到 {employee_name} 的班次。")

        records.sort(key=lambda item: (item.year, item.month, item.day, SHIFT_ORDER[item.shift]))
        return records

    def _locate_cell(
        self,
        box: Sequence[Sequence[float]],
        verticals: Sequence[int],
        horizontals: Sequence[int],
    ) -> Tuple[int, int]:
        xs = [point[0] for point in box]
        ys = [point[1] for point in box]
        center_x = sum(xs) / len(xs)
        center_y = sum(ys) / len(ys)
        cell_col = bisect_right(verticals, center_x) - 1
        cell_row = bisect_right(horizontals, center_y) - 1
        return cell_col, cell_row

    def _normalize_text(self, text: str) -> str:
        return (
            text.strip()
            .replace("（", "(")
            .replace("）", ")")
            .replace(" ", "")
            .replace("月份", "/")
            .replace("排班表", "")
        )

    def _normalize_name(self, text: str) -> str:
        normalized = self._normalize_text(text)
        return NAME_VARIANTS.get(normalized, normalized)


def export_to_excel(records: Sequence[ShiftRecord], employee_name: str, output_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "個人班表"

    headers = ["年", "月", "日", "周幾", "班別", "時數", "時薪", "單量", "總薪水"]
    sheet.append(headers)

    for record in records:
        sheet.append(
            [
                record.year,
                record.month,
                record.day,
                record.weekday,
                record.shift,
                record.hours,
                "",
                "",
                "",
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    title_fill = PatternFill("solid", fgColor="2F6B53")
    header_fill = PatternFill("solid", fgColor="2F6B53")
    border = Border(
        left=Side(style="thin", color="D9E2DC"),
        right=Side(style="thin", color="D9E2DC"),
        top=Side(style="thin", color="D9E2DC"),
        bottom=Side(style="thin", color="D9E2DC"),
    )
    default_row_fill_odd = PatternFill("solid", fgColor="F7FAF8")
    default_row_fill_even = PatternFill("solid", fgColor="EDF4F1")
    base_font = Font(name=FONT_NAME, size=11)
    header_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
    shift_font = Font(name=FONT_NAME, size=11, bold=True, color="2A3B33")

    title = f"{employee_name} 個人班表"
    sheet.insert_rows(1)
    sheet.merge_cells("A1:I1")
    sheet["A1"] = title
    sheet["A1"].font = title_font
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24

    for cell in sheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_index in range(3, sheet.max_row + 1):
        row_fill = default_row_fill_odd if row_index % 2 == 1 else default_row_fill_even
        for cell in sheet[row_index]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = base_font
            cell.fill = row_fill

        shift_cell = sheet[f"E{row_index}"]
        if shift_cell.value in SHIFT_FILLS:
            shift_cell.fill = SHIFT_FILLS[shift_cell.value]
            shift_cell.font = shift_font

    widths = {"A": 10, "B": 8, "C": 8, "D": 8, "E": 10, "F": 8, "G": 12, "H": 12, "I": 14}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    if sheet.max_row >= 3:
        table = Table(displayName="PersonalSchedule", ref=f"A2:I{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium7",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def default_output_path(image_path: Path, employee_name: str) -> Path:
    safe_name = employee_name.replace("/", "_").replace("\\", "_")
    return image_path.with_name(f"{image_path.stem}_{safe_name}_個人班表.xlsx")


def run_conversion(
    image_path: Path,
    employee_name: str,
    output_path: Path | None = None,
) -> Tuple[Path, List[ShiftRecord]]:
    parser = ScheduleParser()
    records = parser.parse(image_path=image_path, employee_name=employee_name)
    final_output = output_path or default_output_path(Path(image_path), employee_name)
    saved_path = export_to_excel(records=records, employee_name=employee_name, output_path=final_output)
    return saved_path, records


def prompt_for_terminal_inputs() -> Tuple[Path, str, Path | None]:
    print("班表圖片轉 Excel")
    print("直接按 Enter 可使用預設值。")

    input_path_text = input("請輸入班表圖片路徑: ").strip().strip('"')
    if not input_path_text:
        raise ValueError("你沒有輸入圖片路徑。")

    employee_name = input("請輸入姓名 [張盈慧]: ").strip() or "張盈慧"
    output_path_text = input("請輸入輸出 Excel 路徑 [直接 Enter 自動命名]: ").strip().strip('"')

    input_path = Path(input_path_text)
    output_path = Path(output_path_text) if output_path_text else None
    return input_path, employee_name, output_path


class ScheduleApp:
    def __init__(self, root: "tk.Tk") -> None:
        self.root = root
        self.root.title("班表圖片轉 Excel")
        self.root.geometry("640x280")
        self.root.resizable(False, False)

        self.image_var = tk.StringVar()
        self.name_var = tk.StringVar(value="張盈慧")
        self.status_var = tk.StringVar(value="請先選擇班表圖片。")

        container = ttk.Frame(root, padding=18)
        container.pack(fill="both", expand=True)

        ttk.Label(container, text="班表圖片").grid(row=0, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(container, textvariable=self.image_var, width=58).grid(row=1, column=0, sticky="ew")
        ttk.Button(container, text="選擇圖片", command=self.pick_image).grid(row=1, column=1, padx=(12, 0))

        ttk.Label(container, text="姓名").grid(row=2, column=0, sticky="w", pady=(18, 8))
        ttk.Entry(container, textvariable=self.name_var, width=24).grid(row=3, column=0, sticky="w")

        ttk.Button(container, text="產生 Excel", command=self.generate_excel).grid(
            row=4, column=0, sticky="w", pady=(22, 10)
        )

        ttk.Label(
            container,
            text="輸出欄位：年 / 月 / 日 / 周幾 / 班別 / 時數 / 時薪 / 單量 / 總薪水",
            foreground="#52615B",
        ).grid(row=5, column=0, columnspan=2, sticky="w", pady=(4, 8))

        ttk.Label(
            container,
            textvariable=self.status_var,
            wraplength=590,
            foreground="#2F4B3F",
        ).grid(row=6, column=0, columnspan=2, sticky="w", pady=(14, 0))

        container.columnconfigure(0, weight=1)

    def pick_image(self) -> None:
        path = filedialog.askopenfilename(
            title="選擇班表圖片",
            filetypes=[("圖片", "*.png;*.jpg;*.jpeg;*.bmp"), ("所有檔案", "*.*")],
        )
        if path:
            self.image_var.set(path)
            self.status_var.set("圖片已選擇，可以直接產生 Excel。")

    def generate_excel(self) -> None:
        image_path = Path(self.image_var.get().strip())
        employee_name = self.name_var.get().strip() or "張盈慧"

        if not image_path.exists():
            messagebox.showerror("找不到圖片", "請先選擇正確的班表圖片。")
            return

        output_path = filedialog.asksaveasfilename(
            title="另存 Excel",
            defaultextension=".xlsx",
            initialfile=default_output_path(image_path, employee_name).name,
            filetypes=[("Excel 檔案", "*.xlsx")],
        )
        if not output_path:
            return

        self.status_var.set("正在辨識班表並產生 Excel，請稍候...")
        self.root.update_idletasks()

        try:
            saved_path, records = run_conversion(image_path, employee_name, Path(output_path))
        except Exception as exc:
            self.status_var.set("處理失敗，請換一張更清晰的班表圖片再試一次。")
            messagebox.showerror("轉換失敗", f"{exc}\n\n{traceback.format_exc()}")
            return

        self.status_var.set(f"完成，共輸出 {len(records)} 筆班次。檔案位置：{saved_path}")
        messagebox.showinfo("完成", f"已成功產生 Excel：\n{saved_path}")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="將班表圖片轉成指定員工的 Excel 個人班表。")
    parser.add_argument("--input", dest="input_path", help="班表圖片路徑")
    parser.add_argument("--name", dest="employee_name", default="張盈慧", help="要擷取的姓名")
    parser.add_argument("--output", dest="output_path", help="輸出的 Excel 路徑")
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    if args.input_path:
        output_path, records = run_conversion(
            image_path=Path(args.input_path),
            employee_name=args.employee_name,
            output_path=Path(args.output_path) if args.output_path else None,
        )
        print(f"已輸出 {len(records)} 筆班次到：{output_path}")
        return

    try:
        image_path, employee_name, output_path = prompt_for_terminal_inputs()
        saved_path, records = run_conversion(
            image_path=image_path,
            employee_name=employee_name,
            output_path=output_path,
        )
        print(f"已輸出 {len(records)} 筆班次到：{saved_path}")
        return
    except EOFError:
        pass
    except KeyboardInterrupt:
        print("\n已取消。")
        return
    except Exception as exc:
        print(f"\n終端機模式失敗：{exc}")
        return

    if tk is None:
        raise RuntimeError("目前環境無法啟動圖形介面，請改用命令列模式。")

    root = tk.Tk()
    style = ttk.Style(root)
    if "clam" in style.theme_names():
        style.theme_use("clam")
    ScheduleApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
